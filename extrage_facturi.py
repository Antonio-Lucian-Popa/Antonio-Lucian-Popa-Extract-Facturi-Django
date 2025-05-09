import os
import re
import fitz  # PyMuPDF
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell


def parse_number(val):
    if not val:
        return 0
    val = val.replace('\xa0', '').replace("−", "-").replace("–", "-").replace("—", "-")
    val = re.sub(r'(?<=\d)\.(?=\d{3}(?:\D|$))', '', val)  # elimină puncte de mii
    val = val.replace(",", ".")
    try:
        return float(val)
    except ValueError:
        return 0


def extract_all_indexes(text, fallback_text=None):
    if fallback_text is None:
        fallback_text = text

    def extract_index_pair(regex):
        match = re.search(regex, text, re.IGNORECASE)
        if not match:
            match = re.search(regex, fallback_text, re.IGNORECASE)
        if match:
            return parse_number(match.group(1)), parse_number(match.group(2))
        return 0, 0

    # Acceptăm "Citire distribuitor" SAU "Estimare convenție" după valorile numerice
    pattern_common = r"(\d{1,3}(?:\.\d{3})*,\d+)\s+(?:Citire distribuitor|Estimare convenție)\s+(\d{1,3}(?:\.\d{3})*,\d+)\s+(?:Citire distribuitor|Estimare convenție)"

    activ_v, activ_n = extract_index_pair(r"Energie activ[ăa].*?" + pattern_common)
    inductiv_v, inductiv_n = extract_index_pair(r"Energie reactiv[ăa] inductiv[ăa].*?" + pattern_common)
    capacitiv_v, capacitiv_n = extract_index_pair(r"Energie reactiv[ăa] capacitiv[ăa].*?" + pattern_common)

    return {
        "index_activ_vechi": activ_v,
        "index_activ_nou": activ_n,
        "index_reactivi_vechi": inductiv_v,
        "index_reactivi_nou": inductiv_n,
        "index_reactivc_vechi": capacitiv_v,
        "index_reactivc_nou": capacitiv_n
    }


def extract_sume_cantitati(text, fallback_text=None):
    if fallback_text is None:
        fallback_text = text

    # Încercăm să izolăm secțiunea „DETALII CITIRI” până la următoarea secțiune
    match_citiri = re.search(r"DETALII CITIRI(.*?)DETALII PRODUSE", fallback_text, re.DOTALL | re.IGNORECASE)
    citiri_text = " ".join(match.group(1) for match in re.finditer(
        r"DETALII CITIRI(.*?)(?=DETALII CITIRI|DETALII PRODUSE|TOTAL|$)", fallback_text, re.DOTALL | re.IGNORECASE))

    match_total = re.search(r"Total loc de consum.*?([\-−–]?\d{1,3}(?:[.,]\d{3})*[.,]?\d+)\s*kWh", fallback_text)


    def suma_cantitati(denumire):
        pattern = rf"{denumire}.*?(?:\d{{2}}\.\d{{2}}\.\d{{4}})?\s*[\d.,]+\s*Citire.*?[\d.,]+\s*Citire.*?([\d.,]+)"
        matches = re.findall(pattern, citiri_text, re.IGNORECASE)
        if not matches:
            matches = re.findall(pattern, fallback_text, re.IGNORECASE)
        return sum(parse_number(v) for v in matches)

    def suma_cantitate_facturata(denumire_fix, x_type="X1", src=None):
        if src is None:
            src = text  # fallback

        # Normalizează
        src = src.replace('\n', ' ').replace('\xa0', ' ').replace('\r', ' ')
        src = re.sub(r'\s+', ' ', src)

        # Construim un regex robust care caută: <denumire> <X1|X3> <dată> <dată> <valoare> kVArh
        denumire_fix_escaped = re.sub(r"\s+", r"\\s+", denumire_fix)
        pattern = rf"{denumire_fix_escaped}\s+{x_type}.*?(?:\d{{2}}\.\d{{2}}\.\d{{2,4}})?\s*-\s*(?:\d{{2}}\.\d{{2}}\.\d{{2,4}})?\s+([\-−–]?\d+(?:[.,]?\d*)?)\s*kVArh"
        matches = re.findall(pattern, src, re.IGNORECASE)

        print(f"[DEBUG] {denumire_fix} {x_type} → {matches}")
        return sum(parse_number(v) for v in matches)

    # === nou: căutare energie reactivă X1 și X3 separat pentru capacitiv/inductiv ===
    def suma_reactivi():
        return suma_cantitate_facturata("Energie reactiv[ăa] inductiv[ăa] X1")

    def suma_reactivc():
        return (
            suma_cantitate_facturata("Energie reactiv[ăa] capacitiv[ăa] X1") +
            suma_cantitate_facturata("Energie reactiv[ăa] capacitiv[ăa] X3")
        )

    return {
        "cantitate_activ": suma_cantitati("Energie activ[ăa]"),
        "cantitate_reactivi": suma_cantitati("Energie reactiv[ăa] inductiv[ăa]"),
        "cantitate_reactivc": suma_cantitati("Energie reactiv[ăa] capacitiv[ăa]"),
        "cantitate_facturata_activ": parse_number(match_total.group(1)) if match_total else 0,
        "cantitate_facturata_reactivi": (
            suma_cantitate_facturata("Energie reactivă inductivă", "X1", text) +
            suma_cantitate_facturata("Energie reactivă inductivă", "X3", text)
        ),
        "cantitate_facturata_reactivc": (
            suma_cantitate_facturata("Energie reactivă capacitivă", "X1", text) +
            suma_cantitate_facturata("Energie reactivă capacitivă", "X3", text)
        ),


    }


def extract_data_from_text(text, global_text=None):
    if global_text is None:
        global_text = text
        # Procesăm separat textul local (doar blocul) și textul global (întregul fișier)
    local_text = text.replace('\n', ' ').replace('\r', '').replace('\xa0', ' ')
    full_text = global_text.replace('\n', ' ').replace('\r', '').replace('\xa0', ' ')

    # 👉 1. Extragem loc_consum DOAR din bloc (nu din global_text)
    zona_consum = ""
    match_consum_section = re.search(
        r"DETALII LOC DE (?:CONSUM|PRODUCERE ȘI CONSUM)[\s\-–—:]*?(.*?)(?:Denumirea produsului\s*contractat|COD Loc de consum)",
        local_text,
        flags=re.DOTALL | re.IGNORECASE
    )
    if match_consum_section:
        zona_consum = match_consum_section.group(1).strip()
    else:
        alt_match = re.search(
            r"((?:Localitatea|Comuna)[^:]*?Cod postal\s+\d{5,6})\s+Denumirea produsului\s*contractat",
            local_text,
            re.IGNORECASE
        )
        if alt_match:
            zona_consum = alt_match.group(1).strip()

    loc_consum_match = re.search(
        r"(?:Localitatea|Comuna)\s+[A-ZȘȚĂÎÂ].*?Cod postal\s+\d{5,6}",
        zona_consum,
        flags=re.IGNORECASE
    )

    def find(pattern, src=None, group=1):
        if src is None:
            src = local_text
        match = re.search(pattern, src, re.DOTALL | re.IGNORECASE)
        if match:
            try:
                return match.group(group).strip()
            except IndexError:
                print(f"[‼️] Regex fără grupul {group}: {pattern}")
                return ""
        return ""

    data = {
        # date care sunt DOAR în bloc
        "loc_consum": loc_consum_match.group(0).strip() if loc_consum_match else "",
        "POD": find(r"POD:?\s*([A-Z0-9]{8,})", text),

        # date care sunt doar în header / prima pagină → extragem din global_text
        "factura": find(r"(?:Nr\. factura|Serie\s*/\s*Nr\.):?\s*([A-Z]+/?\d+)", global_text),
        "data_emitere": find(r"Dat[ăa] emitere:?\s*(\d{2}\.\d{2}\.\d{4})", global_text),
        "data_scadenta": find(r"Data scadent[ăa]:?\s*(\d{2}\.\d{2}\.\d{4})", global_text),
        "perioada_start": find(r"Perioad[ăa] (?:de facturare)?:?\s*(\d{2}\.\d{2}\.\d{4})", global_text),
        "perioada_end": find(r"Perioad[ăa] (?:de facturare)?:?\s*\d{2}\.\d{2}\.\d{4} - (\d{2}\.\d{2}\.\d{4})",
                             global_text),
        "total_net": parse_number(
            find(r"Valoare facturat[ăa] f[ăa]r[ăa] TVA.*?([\-−–]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))", global_text)),
        "valoare_fara_TVA": parse_number(
            find(r"Valoare facturat[ăa] f[ăa]r[ăa] TVA.*?([\-−–]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))", global_text)),
        "valoare_cu_TVA": parse_number(
            find(r"TOTAL FACTUR[ĂA] CURENT[ĂA] CU TVA\s+([\-−–]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))", global_text)),
        "total_plata": parse_number(
            find(r"TOTAL DE PLAT[ĂA][^\d\-]*([\-−–]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))", global_text) or find(
                r"Cod de bare.*?([\-−–]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))", global_text)),
        "sold_anterior": parse_number(
            find(r"Sold la data emiterii facturii.*?([\-−–]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))", global_text)),

        # opționale, din bloc (dacă există)
        "index_vechi": parse_number(find(r"Index vechi[^0-9]*([\d.,]+)", text)),
        "index_nou": parse_number(find(r"Index nou[^0-9]*([\d.,]+)", text)),
        "cantitate": parse_number(
            find(r"Total EA.*?([\d.,]+)\s*kWh", text) or
            find(r"Cantitate facturat[ăa]\s*([\d.,]+)\s*kWh", text) or
            find(r"Total energie activ[ăa]\s*([\d.,]+)\s*kWh", text))
    }

    data.update(extract_all_indexes(local_text, fallback_text=full_text))
    data.update(extract_sume_cantitati(local_text, fallback_text=full_text))

    return data


def split_pdf_by_blocuri(text):
    blocuri = re.split(r"(?=DETALII LOC DE (?:CONSUM|PRODUCERE ȘI CONSUM))", text, flags=re.IGNORECASE)

    rezultate = []
    for bloc in blocuri:
        upper_bloc = bloc.upper()
        if "POD" in upper_bloc and "LOCALITATEA" in upper_bloc:
            rezultate.append(bloc)
    return rezultate



def process_pdfs(folder_path, output_excel_path):
    rows = []
    for filename in os.listdir(folder_path):
        if not filename.lower().endswith(".pdf"):
            continue
        pdf_path = os.path.join(folder_path, filename)
        try:
            doc = fitz.open(pdf_path)
            full_text = "\n".join(page.get_text() for page in doc)
            blocuri = split_pdf_by_blocuri(full_text)
            for bloc in blocuri:
                data = extract_data_from_text(bloc, global_text=full_text)

                # ✅ Calcul diferențe
                delta_activ = round(data["index_activ_nou"] - data["index_activ_vechi"], 3)
                delta_reactivi = round(data["index_reactivi_nou"] - data["index_reactivi_vechi"], 3)
                delta_reactivc = round(data["index_reactivc_nou"] - data["index_reactivc_vechi"], 3)

                alert = []
                if abs(delta_activ - data["cantitate_activ"]) > 1:
                    alert.append("index activ ≠ cantitate")

                if abs(delta_reactivi - data["cantitate_reactivi"]) > 1:
                    alert.append("index reactiv I ≠ cantitate")

                if abs(delta_reactivc - data["cantitate_reactivc"]) > 1:
                    alert.append("index reactiv C ≠ cantitate")

                data["alerta"] = " | ".join(alert)

                if any([v != 0 for k, v in data.items() if k.startswith("index_") or k.startswith("cantitate")]):
                    data["fisier"] = filename
                    rows.append(data)

            print(f"[✔] Procesat: {filename} ({len(blocuri)} blocuri)")
        except Exception as e:
            print(f"[⚠️] Eroare la {filename}: {e}")
    return rows


def finalize_excel(data_rows, output_excel_path):
    df = pd.DataFrame(data_rows)

    multi_columns = pd.MultiIndex.from_tuples([
        ("", "loc consum"), ("", "PERIOADA CONSUM"), ("", ""), ("", "POD"), ("", "factura"),
        ("", "pret"), ("", "valoare_fara_TVA"), ("", "valoare_cu_TVA"), ("", "total_plata"),
        ("", "sold_anterior"), ("index vechi", "activ"), ("index vechi", "reactiv I"), ("index vechi", "reactiv C"),
        ("index nou", "activ"), ("index nou", "reactiv I"), ("index nou", "reactiv C"),
        ("cantitate citita", "activ"), ("cantitate citita", "reactiv I"), ("cantitate citita", "reactiv C"),
        ("cantitate facturata", "activ"), ("cantitate facturata", "reactiv I"), ("cantitate facturata", "reactiv C"),
        ("", "fisier"), ("", "alerta")
    ])

    df_reordered = pd.DataFrame([
        [
            row.get("loc_consum", ""),
            row.get("perioada_start", "") + "-" + row.get("perioada_end", ""),
            "",
            row.get("POD", ""),
            row.get("factura", ""),
            row.get("total_net", 0),
            row.get("valoare_fara_TVA", 0),
            row.get("valoare_cu_TVA", 0),
            row.get("total_plata", 0),
            row.get("sold_anterior", 0),
            row.get("index_activ_vechi", 0),
            row.get("index_reactivi_vechi", 0),
            row.get("index_reactivc_vechi", 0),
            row.get("index_activ_nou", 0),
            row.get("index_reactivi_nou", 0),
            row.get("index_reactivc_nou", 0),
            row.get("cantitate_activ", 0),
            row.get("cantitate_reactivi", 0),
            row.get("cantitate_reactivc", 0),
            row.get("cantitate_facturata_activ", 0),
            row.get("cantitate_facturata_reactivi", 0),
            row.get("cantitate_facturata_reactivc", 0),
            row.get("fisier", ""),
            "⚠️ Diferență mare la reactiv C" if abs(
                row.get("cantitate_reactivc", 0) - row.get("cantitate_facturata_reactivc", 0)) > 100 else ""
        ] for row in data_rows
    ], columns=multi_columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturi"
    ws.freeze_panes = "A3"

    for r in dataframe_to_rows(df_reordered, index=False, header=True):
        ws.append(r)

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = bold_center
            cell.alignment = center
            cell.border = thin_border

    merge_groups = {
        "index vechi": (11, 13),
        "index nou": (14, 16),
        "cantitate citita": (17, 19),
        "cantitate facturata": (20, 22)
    }
    for title, (start_col, end_col) in merge_groups.items():
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).value = title
        for col in range(start_col, end_col + 1):
            ws.cell(row=2, column=col).value = multi_columns[col - 1][1]

    fill_colors = {
        "index_vechi": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "index_nou": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
        "cantitate_citita": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "cantitate_facturata": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "cantitate_facturata_c": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    }

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.border = thin_border
            col = cell.column
            if col == 24 and cell.value:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif col in [11, 12, 13]:
                cell.fill = fill_colors["index_vechi"]
            elif col in [14, 15, 16]:
                cell.fill = fill_colors["index_nou"]
            elif col in [17, 18, 19]:
                cell.fill = fill_colors["cantitate_citita"]
            elif col == 20:
                cell.fill = fill_colors["cantitate_facturata"]
            elif col in [21, 22]:
                cell.fill = fill_colors["cantitate_facturata_c"]

    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if not isinstance(cell, MergedCell):
                col_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_excel_path)
